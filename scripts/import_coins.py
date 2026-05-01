#!/usr/bin/env python3
"""
Import a 'Coin Collection.xlsx' into Roundhouse's coin collection.

Reads an .xlsx with the columns:
    Type | Country | Currency | Denomination | Year | Mint | Condition |
    Quantity | Value | Total | Comment

and inserts every row into the kind='coins' collection. The collection
is auto-created by the migration runner; this script just populates it.

The "Currency" column in the spreadsheet is misleadingly named — it's
actually the numeric face value (5, 10, 20). The "Denomination" column
holds the currency unit name (Dollar, Pesos, Yuan).

Idempotent guard: refuses to run if the coin collection already has
items. Pass --force to wipe-and-reimport.

Usage:
    python3 scripts/import_coins.py [path/to/Coin Collection.xlsx] [--force]
"""

from __future__ import annotations

import argparse
import sqlite3
import sys
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

DEFAULT_XLSX = Path.home() / "Downloads" / "Coin Collection.xlsx"
DEFAULT_DB = Path.home() / "Library" / "Application Support" / "roundhouse" / "roundhouse.db"


def clean(v: object) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    if not s or s in ("?", "-"):
        return None
    return s


def to_float(v: object) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def to_int(v: object, default: Optional[int] = None) -> Optional[int]:
    if v is None:
        return default
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v)
    s = str(v).strip()
    if not s:
        return default
    try:
        return int(float(s))
    except ValueError:
        return default


def to_cents(v: object) -> Optional[int]:
    f = to_float(v)
    if f is None:
        return None
    return int(round(f * 100))


def map_type(raw: object) -> str:
    s = (clean(raw) or "").lower()
    if s.startswith("bill"):
        return "bill"
    if s.startswith("coin"):
        return "coin"
    return "coin"


def synthesize_name(country: Optional[str], face_value: Optional[float],
                    denomination: Optional[str], year: Optional[int],
                    mint: Optional[str]) -> str:
    parts: list[str] = []
    if year is not None:
        parts.append(str(year))
    if country:
        parts.append(country)
    if face_value is not None:
        # Render whole numbers without trailing .0
        if face_value == int(face_value):
            parts.append(str(int(face_value)))
        else:
            parts.append(str(face_value))
    if denomination:
        parts.append(denomination)
    if mint:
        parts.append(f"({mint})")
    return " ".join(parts) or "(unnamed coin)"


def main() -> int:
    ap = argparse.ArgumentParser(description="Import a coin/bill .xlsx into Roundhouse")
    ap.add_argument("xlsx", nargs="?", default=str(DEFAULT_XLSX),
                    help="Path to the .xlsx file (default: ~/Downloads/Coin Collection.xlsx)")
    ap.add_argument("--db", default=str(DEFAULT_DB),
                    help="Target SQLite DB (default: Roundhouse user data dir)")
    ap.add_argument("--force", action="store_true",
                    help="Delete the coin collection's items first if any exist")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx).expanduser()
    db_path = Path(args.db).expanduser()

    if not xlsx_path.exists():
        print(f"❌ Spreadsheet not found: {xlsx_path}", file=sys.stderr)
        return 2
    if not db_path.exists():
        print(f"❌ DB not found at {db_path}.", file=sys.stderr)
        print("   Launch Roundhouse once to create it (the migration runner will create the Coin Collection automatically), then rerun.",
              file=sys.stderr)
        return 2

    print(f"📖 Reading {xlsx_path}")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # rows[0] = header. Filter to rows that actually look like data.
    data_rows = []
    for r in rows[1:]:
        if not r:
            continue
        # A real row has either a Type or Country populated.
        if not (clean(r[0] if len(r) > 0 else None) or clean(r[1] if len(r) > 1 else None)):
            continue
        data_rows.append(r)
    print(f"   {len(data_rows)} data rows found")

    db = sqlite3.connect(db_path)
    db.execute("PRAGMA foreign_keys = ON")

    # Sanity-check the schema is v0.5.0 (kind-aware).
    coll_cols = [r[1] for r in db.execute("PRAGMA table_info(collections)").fetchall()]
    if "kind" not in coll_cols:
        print("❌ collections.kind column missing. Launch v0.5.0+ of Roundhouse first to migrate the DB.", file=sys.stderr)
        return 3

    coin_coll = db.execute(
        "SELECT id FROM collections WHERE kind = 'coins' ORDER BY id LIMIT 1"
    ).fetchone()
    if not coin_coll:
        print("⚠  No kind='coins' collection found. Launch Roundhouse once so the migration creates it, then rerun.", file=sys.stderr)
        return 4
    coin_coll_id = coin_coll[0]
    print(f"   target collection_id={coin_coll_id}")

    existing_count = db.execute(
        "SELECT COUNT(*) FROM items WHERE collection_id = ?", (coin_coll_id,)
    ).fetchone()[0]
    if existing_count and not args.force:
        print(f"⚠  Coin collection already has {existing_count} items. Use --force to wipe and reimport.",
              file=sys.stderr)
        return 5
    if existing_count and args.force:
        print(f"   --force: deleting {existing_count} existing items in coin collection")
        db.execute("DELETE FROM items WHERE collection_id = ?", (coin_coll_id,))

    insert_sql = """
        INSERT INTO items (
            collection_id, type, name,
            country, face_value, denomination, mint_mark, quantity,
            year, condition,
            current_value_cents, notes
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """

    inserted = 0
    by_country: dict[str, int] = {}
    by_type: dict[str, int] = {}

    with db:
        for r in data_rows:
            r = list(r) + [None] * (12 - len(r))
            type_raw, country_raw, currency_raw, denom_raw, year_raw, mint_raw, \
                condition_raw, qty_raw, value_raw, _total, comment_raw, _ = r[:12]

            country = clean(country_raw)
            face_value = to_float(currency_raw)
            denomination = clean(denom_raw)
            year = to_int(year_raw)
            mint = clean(mint_raw)
            qty = to_int(qty_raw, 1) or 1
            condition = clean(condition_raw)
            value_cents = to_cents(value_raw)
            notes = clean(comment_raw)
            t = map_type(type_raw)
            name = synthesize_name(country, face_value, denomination, year, mint)

            db.execute(insert_sql, (
                coin_coll_id, t, name,
                country, face_value, denomination, mint, qty,
                year, condition,
                value_cents, notes
            ))
            inserted += 1
            if country:
                by_country[country] = by_country.get(country, 0) + 1
            by_type[t] = by_type.get(t, 0) + 1

    print(f"\n✓ Inserted {inserted} record(s) into the coin collection")
    print("\n  by type:")
    for t, n in sorted(by_type.items(), key=lambda x: -x[1]):
        print(f"    {t:>10} : {n}")
    print(f"\n  countries: {len(by_country)}")
    for c, n in sorted(by_country.items(), key=lambda x: -x[1])[:10]:
        print(f"    {c:>20} : {n}")
    if len(by_country) > 10:
        print(f"    ... ({len(by_country) - 10} more)")

    total_value = db.execute(
        "SELECT COALESCE(SUM(current_value_cents * quantity), 0) FROM items WHERE collection_id = ?",
        (coin_coll_id,)
    ).fetchone()[0]
    print(f"\n  total current value: ${total_value / 100:,.2f}")

    db.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
