#!/usr/bin/env python3
"""
Bulk-import a model-train inventory spreadsheet into Roundhouse.

Reads an .xlsx with the columns:
    Scale | Mfg | Number | Item | Source | Purchased | Price | Color | Notes

and inserts every row as an item under a single collection. Items are not
assigned to a set; the user can group them in the UI later.

The category prefix on the 'Item' column ("Locomotive - …", "RS - …",
"Track - …", etc.) is mapped to the Roundhouse `type` enum.

Idempotent guard: refuses to run if the target collection already exists.
Pass --force to wipe-and-reimport that collection.

Usage:
    python3 scripts/import_xlsx.py path/to/inventory.xlsx \\
        [--collection NAME] [--description TEXT] [--force]
"""

from __future__ import annotations

import argparse
import sqlite3
import sys
from datetime import datetime, date
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

# Mac dev path. On Windows this script wouldn't run anyway since Roundhouse
# isn't packaged for Mac use long-term — this is dev-only seed data.
DEFAULT_DB = Path.home() / "Library" / "Application Support" / "roundhouse" / "roundhouse.db"
DEFAULT_COLLECTION_NAME = "Carey's Trains"

VALID_SCALES = {"Z", "N", "HO", "OO", "S", "O", "G"}


def map_scale(raw: object) -> Optional[str]:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    up = s.upper()
    if up in VALID_SCALES:
        return up
    if up in ("N/A", "ALL", "ANY"):
        return None
    # 1/43, 1/64, ON30, etc — keep but bucket as 'other'
    return "other"


def map_type(item_name: str) -> str:
    if not item_name:
        return "other"
    s = item_name.strip()
    # The category prefix is usually before " - "
    prefix = s.split(" - ", 1)[0].strip().lower()

    if prefix == "rs" or prefix.startswith("rs "):
        return "rolling_stock"
    if prefix.startswith("locomotive"):
        return "locomotive"
    if prefix.startswith("track"):
        return "track"
    if prefix.startswith("accessor"):
        return "accessory"
    if prefix.startswith("building") or "structure" in prefix or "sturcture" in prefix:
        return "building"
    if prefix in ("toy truck", "truck", "ahl", "hartoy"):
        return "accessory"
    if prefix.startswith("transformer") or prefix.startswith("tools"):
        return "accessory"
    if prefix.startswith("trolley"):
        return "rolling_stock"
    # explicit "other" buckets
    if prefix.startswith(
        (
            "puzzle",
            "coin bank",
            "train set",
            "book",
            "repairs",
            "lot",
            "misc",
        )
    ):
        return "other"
    return "other"


def to_iso_date(v: object) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    return s or None


def to_cents(v: object) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(round(v * 100))
    s = str(v).strip().replace("$", "").replace(",", "")
    if not s:
        return None
    try:
        return int(round(float(s) * 100))
    except ValueError:
        return None


def clean(v: object) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    if not s or s in ("?", "-"):
        return None
    return s


def main() -> int:
    ap = argparse.ArgumentParser(description="Import a train-inventory .xlsx into Roundhouse")
    ap.add_argument("xlsx", help="Path to the .xlsx inventory file")
    ap.add_argument("--db", default=str(DEFAULT_DB),
                    help="Target SQLite DB (default: Roundhouse user data dir)")
    ap.add_argument("--collection", default=DEFAULT_COLLECTION_NAME,
                    help=f"Collection name to import into (default: {DEFAULT_COLLECTION_NAME!r})")
    ap.add_argument("--description", default=None,
                    help="Collection description (default: 'Imported from <filename>')")
    ap.add_argument("--force", action="store_true",
                    help="Delete the target collection before importing if it exists")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx).expanduser()
    db_path = Path(args.db).expanduser()
    collection_name: str = args.collection
    collection_desc: str = args.description or f"Imported from {xlsx_path.name}"

    if not xlsx_path.exists():
        print(f"❌ Spreadsheet not found: {xlsx_path}", file=sys.stderr)
        return 2

    if not db_path.exists():
        # Bootstrap the DB by applying the project schema directly.
        schema_path = Path(__file__).resolve().parent.parent / "src" / "main" / "db" / "schema.sql"
        if not schema_path.exists():
            print(f"❌ DB missing and schema.sql not found at {schema_path}", file=sys.stderr)
            return 2
        print(f"🆕 Creating DB at {db_path}")
        db_path.parent.mkdir(parents=True, exist_ok=True)
        boot = sqlite3.connect(db_path)
        boot.executescript(schema_path.read_text())
        boot.execute("PRAGMA journal_mode = WAL")
        boot.commit()
        boot.close()

    print(f"📖 Reading {xlsx_path}")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # rows[0] = header, rows[1] = totals row, rows[2+] = data
    # but we tolerate either by filtering for an actual item name.
    data_rows = [r for r in rows[1:] if r[3] and isinstance(r[3], str) and r[3].strip()]
    print(f"   {len(data_rows)} item rows found")

    db = sqlite3.connect(db_path)
    db.execute("PRAGMA foreign_keys = ON")

    # Sanity-check the schema has the source column.
    cols = [r[1] for r in db.execute("PRAGMA table_info(items)").fetchall()]
    if "source" not in cols:
        print("❌ items.source column missing. Wipe the DB and relaunch the app to apply the new schema.",
              file=sys.stderr)
        return 3

    # Sanity-check we have the kind column (v0.5.0+).
    coll_cols = [r[1] for r in db.execute("PRAGMA table_info(collections)").fetchall()]
    has_kind = "kind" in coll_cols
    has_collection_id = "collection_id" in cols

    existing = db.execute("SELECT id FROM collections WHERE name = ?",
                          (collection_name,)).fetchone()
    if existing and not args.force:
        print(f"⚠  Collection '{collection_name}' already exists. Use --force to wipe and reimport.",
              file=sys.stderr)
        return 4
    if existing and args.force:
        print(f"   --force: deleting existing collection (id={existing[0]}) and its sets/items")
        db.execute("DELETE FROM collections WHERE id = ?", (existing[0],))

    if has_kind:
        cur = db.execute(
            "INSERT INTO collections (name, description, kind) VALUES (?, ?, 'trains')",
            (collection_name, collection_desc),
        )
    else:
        cur = db.execute(
            "INSERT INTO collections (name, description) VALUES (?, ?)",
            (collection_name, collection_desc),
        )
    coll_id = cur.lastrowid
    print(f"✓ Created collection '{collection_name}' (id={coll_id})")

    # Counters for the report
    by_type: dict[str, int] = {}
    by_scale: dict[str, int] = {}
    inserted = 0
    skipped = 0

    if has_collection_id:
        insert_sql = """
            INSERT INTO items (
                collection_id, set_id, type, name, manufacturer, model_number, scale,
                road_name, era, year, condition, original_box,
                purchase_date, purchase_price_cents, current_value_cents,
                storage_location, source, notes
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
    else:
        insert_sql = """
            INSERT INTO items (
                set_id, type, name, manufacturer, model_number, scale,
                road_name, era, year, condition, original_box,
                purchase_date, purchase_price_cents, current_value_cents,
                storage_location, source, notes
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """

    with db:
        for r in data_rows:
            scale_raw, mfg, number, item_name, source, purchased, price, color, notes = r[:9]

            name = clean(item_name)
            if not name:
                skipped += 1
                continue

            type_ = map_type(name)
            scale = map_scale(scale_raw)

            # Color goes into notes if present, since we don't have a color column.
            note_parts = []
            if clean(color):
                note_parts.append(f"Color: {clean(color)}")
            if clean(notes):
                note_parts.append(str(clean(notes)))
            merged_notes = " · ".join(note_parts) if note_parts else None

            base_row = (
                None,                          # set_id
                type_,                         # type
                name,                          # name
                clean(mfg),                    # manufacturer
                clean(number),                 # model_number
                scale,                         # scale
                None,                          # road_name (not in source)
                None,                          # era
                None,                          # year
                None,                          # condition
                None,                          # original_box
                to_iso_date(purchased),        # purchase_date
                to_cents(price),               # purchase_price_cents
                None,                          # current_value_cents
                None,                          # storage_location
                clean(source),                 # source
                merged_notes,                  # notes
            )
            row = (coll_id, *base_row) if has_collection_id else base_row
            db.execute(insert_sql, row)
            inserted += 1
            by_type[type_] = by_type.get(type_, 0) + 1
            if scale:
                by_scale[scale] = by_scale.get(scale, 0) + 1

    print(f"\n✓ Inserted {inserted} item(s); skipped {skipped}")
    print("\n  by type:")
    for t, n in sorted(by_type.items(), key=lambda x: -x[1]):
        print(f"    {t:>14} : {n}")
    print("\n  by scale:")
    for s, n in sorted(by_scale.items(), key=lambda x: -x[1]):
        print(f"    {s:>14} : {n}")

    total = db.execute(
        "SELECT COALESCE(SUM(purchase_price_cents),0) FROM items WHERE id IN "
        "(SELECT id FROM items)"
    ).fetchone()[0]
    print(f"\n  total purchase value: ${total / 100:,.2f}")

    db.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
